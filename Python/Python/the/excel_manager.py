import openpyxl
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime
import logging

logger = logging.getLogger("ExcelManager")

class ExcelManager:
    FILE_NAME = "Trading_Analytics.xlsx"
    MARKET_STATE_FILE = "market_state.xlsx"
    SIGNAL_ANALYSIS_FILE = "signal_analysis.xlsx"
    PAPER_TRADES_FILE = "paper_trades.xlsx"
    RISK_DRAWDOWN_FILE = "risk_and_drawdown.xlsx"
    POST_MARKET_LEARNING_FILE = "post_market_learning.xlsx"
    
    FILES = {
        MARKET_STATE_FILE: {
            "Market_State": ["Timestamp", "Market_Status", "Symbol", "Candle_Freshness", "Regime"]
        },
        SIGNAL_ANALYSIS_FILE: {
            "Signal_Analysis": ["Timestamp", "Symbol", "Indicator_Values", "Signal_Strength", "Final_Decision", "No_Trade_Reason"]
        },
        PAPER_TRADES_FILE: {
            "Paper_Trades": ["Trade_ID", "Entry_Price", "Exit_Price", "Fake_Capital", "Leverage", "SL", "TP", "PnL", "Exit_Reason"]
        },
        RISK_DRAWDOWN_FILE: {
            "Risk_Drawdown": ["Timestamp", "Current_Equity", "Peak_Equity", "Drawdown_Pct", "Risk_Per_Trade", "Rule_Violations"]
        },
        POST_MARKET_LEARNING_FILE: {
            "Learning": ["Timestamp", "Strategy_Mistake", "Market_Misread_Reason", "Confidence_Mismatch", "Suggested_Improvement"]
        }
    }

    def __init__(self):
        for file_name, sheets in self.FILES.items():
            if not os.path.exists(file_name):
                self._create_specific_workbook(file_name, sheets)

    def _create_specific_workbook(self, file_name, sheets):
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)
        for sheet_name, columns in sheets.items():
            ws = wb.create_sheet(sheet_name)
            ws.append(columns)
        wb.save(file_name)
        logger.info(f"Created new Excel workbook: {file_name}")

    def append_to_file(self, file_name, sheet_name, data_dict):
        try:
            if not os.path.exists(file_name):
                self._create_specific_workbook(file_name, self.FILES[file_name])
            
            wb = load_workbook(file_name)
            ws = wb[sheet_name]
            columns = self.FILES[file_name][sheet_name]
            row = [data_dict.get(col, "") for col in columns]
            ws.append(row)
            wb.save(file_name)
        except Exception as e:
            logger.error(f"Error appending to {file_name}: {e}")

    def update_trade_exit(self, trade_id, exit_data):
        try:
            wb = load_workbook(self.FILE_NAME)
            ws = wb["Trade_Log"]
            
            # Find the row with trade_id
            target_row = None
            for row in ws.iter_rows(min_row=2):
                if row[0].value == trade_id:
                    target_row = row[0].row
                    break
            
            if target_row:
                columns = self.SHEETS["Trade_Log"]
                for key, value in exit_data.items():
                    if key in columns:
                        col_idx = columns.index(key) + 1
                        ws.cell(row=target_row, column=col_idx).value = value
                wb.save(self.FILE_NAME)
            else:
                logger.warning(f"Trade ID {trade_id} not found in Excel for update.")
        except Exception as e:
            logger.error(f"Error updating Excel trade {trade_id}: {e}")

excel_manager = ExcelManager()
